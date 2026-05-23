"""
Generate a formatted Excel reference document for COT report columns.
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "COT Column Reference"

# ── Colour palette ──────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
GOLD        = "C9A227"
LIGHT_GOLD  = "FFF2CC"
GREY_HEADER = "404040"
WHITE       = "FFFFFF"
LIGHT_GREY  = "F2F2F2"

# Section header colours
SECTION_COLOURS = {
    "Identifiers":                     ("1F4E79", "DEEBF7"),
    "Open Interest":                   ("1F4E79", "DEEBF7"),
    "Non-Commercial Positions":        ("375623", "E2EFDA"),
    "Commercial Positions":            ("843C0C", "FCE4D6"),
    "Total Reportable Positions":      ("4F3683", "EDE7F6"),
    "Non-Reportable Positions":        ("4C4C4C", "EDEDED"),
    "Percentage of Open Interest":     ("7B5E00", "FFFBDE"),
    "Number of Traders":               ("1B5E6D", "D0EBF0"),
    "Concentration Ratios":            ("6B2737", "F4D9DF"),
    "Week-over-Week Changes":          ("3B3B00", "F5F5DC"),
}

# ── Helper styles ────────────────────────────────────────────────────────────
thin = Side(style="thin", color="BFBFBF")
medium = Side(style="medium", color="8E8E8E")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

def header_font(color=WHITE, size=11, bold=True):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def cell_font(bold=False, size=10, color="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def wrap_align(horizontal="left"):
    return Alignment(wrap_text=True, vertical="top", horizontal=horizontal)

# ── Title row ────────────────────────────────────────────────────────────────
ws.merge_cells("A1:E1")
title_cell = ws["A1"]
title_cell.value = "CFTC Commitment of Traders — Column Reference (S&P 500 Futures / CME)"
title_cell.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
title_cell.fill = fill(DARK_BLUE)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:E2")
sub_cell = ws["A2"]
sub_cell.value = (
    "Legacy Futures-Only format  |  Source: CFTC  |  "
    "Each row is a weekly Tuesday snapshot released on Friday"
)
sub_cell.font = Font(name="Calibri", italic=True, size=10, color=WHITE)
sub_cell.fill = fill(MID_BLUE)
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

# ── Column-header row ────────────────────────────────────────────────────────
col_headers = ["Column Name", "Data Type", "Section", "Short Label", "Full Description"]
for col_idx, hdr in enumerate(col_headers, start=1):
    cell = ws.cell(row=3, column=col_idx, value=hdr)
    cell.font = header_font(size=10)
    cell.fill = fill(GREY_HEADER)
    cell.alignment = wrap_align("center")
    cell.border = thin_border
ws.row_dimensions[3].height = 20

# ── Data definitions ─────────────────────────────────────────────────────────
# (column_name, dtype, section, short_label, description)
rows = [
    # Identifiers
    ("report_week",      "str",     "Identifiers", "Report Date",
     "Date of the Tuesday snapshot. COT data is always as-of close on Tuesday and published the following Friday."),
    ("contract_units",   "str",     "Identifiers", "Contract Unit",
     "The denomination of one contract (e.g. '$250 × S&P 500 Index')."),
    ("id",               "str",     "Identifiers", "CFTC Contract ID",
     "CFTC internal identifier for the contract."),

    # Open Interest
    ("open_interest_all",  "int64", "Open Interest", "OI — All Months",
     "Total number of open futures contracts across ALL delivery months. "
     "Each contract has one long and one short side — open interest counts only one side. "
     "It equals total long positions (or total short positions) in the market."),
    ("open_interest_old",  "int64", "Open Interest", "OI — Front Month",
     "Open interest in the nearest/front-month (nearby) contract only."),

    # Non-Commercial
    ("noncomm_positions_long_all",   "int64", "Non-Commercial Positions", "Spec Long — All",
     "Gross LONG contracts held by non-commercial (speculative) traders across all delivery months. "
     "Non-commercials have no hedging purpose; they trade for directional profit (e.g. hedge funds, CTAs, retail above reporting threshold)."),
    ("noncomm_positions_short_all",  "int64", "Non-Commercial Positions", "Spec Short — All",
     "Gross SHORT contracts held by non-commercial traders across all delivery months."),
    ("noncomm_postions_spread_all",  "int64", "Non-Commercial Positions", "Spec Spread — All",
     "SPREAD positions of non-commercial traders across all months. "
     "A spread means the trader holds both a long in one delivery month AND a short in a different delivery month simultaneously. "
     "These offsetting contracts are separated from the directional long/short counts above."),
    ("noncomm_positions_long_old",   "int64", "Non-Commercial Positions", "Spec Long — Front Month",
     "Non-commercial gross LONG positions in the front/nearby contract month only."),
    ("noncomm_positions_short_old",  "int64", "Non-Commercial Positions", "Spec Short — Front Month",
     "Non-commercial gross SHORT positions in the front/nearby contract month only."),
    ("noncomm_positions_spread",     "int64", "Non-Commercial Positions", "Spec Spread — Front Month",
     "Non-commercial spread positions in the front/nearby contract month only."),

    # Commercial
    ("comm_positions_long_all",   "int64", "Commercial Positions", "Hedger Long — All",
     "Gross LONG contracts held by commercial (hedging) traders across all delivery months. "
     "Commercials have registered a hedging exemption with the CFTC — they use futures to offset risk in their underlying business "
     "(e.g. an institution hedging equity portfolio exposure)."),
    ("comm_positions_short_all",  "int64", "Commercial Positions", "Hedger Short — All",
     "Gross SHORT contracts held by commercial (hedging) traders across all delivery months."),
    ("comm_positions_long_old",   "int64", "Commercial Positions", "Hedger Long — Front Month",
     "Commercial gross LONG positions in the front/nearby contract month only."),
    ("comm_positions_short_old",  "int64", "Commercial Positions", "Hedger Short — Front Month",
     "Commercial gross SHORT positions in the front/nearby contract month only."),

    # Total Reportable
    ("tot_rept_positions_long_all",  "int64", "Total Reportable Positions", "Total Rept Long — All",
     "Total LONG contracts held by ALL reportable traders (non-commercial + commercial combined) across all months. "
     "'Reportable' means the trader holds positions at or above the CFTC's minimum reporting threshold."),
    ("tot_rept_positions_short",     "int64", "Total Reportable Positions", "Total Rept Short — All",
     "Total SHORT contracts held by all reportable traders across all months."),
    ("tot_rept_positions_long_old",  "int64", "Total Reportable Positions", "Total Rept Long — Front Month",
     "Total reportable LONG positions in the front/nearby contract month only."),
    ("tot_rept_positions_short_1",   "int64", "Total Reportable Positions", "Total Rept Short — Front Month",
     "Total reportable SHORT positions in the front/nearby contract month only. "
     "(The '_1' suffix is a naming-collision artifact in the source data.)"),

    # Non-Reportable
    ("nonrept_positions_long_all",   "int64", "Non-Reportable Positions", "Small Trader Long — All",
     "Estimated LONG contracts held by traders BELOW the CFTC reporting threshold (small traders). "
     "This is a derived figure: Open Interest minus Total Reportable Longs."),
    ("nonrept_positions_short_all",  "int64", "Non-Reportable Positions", "Small Trader Short — All",
     "Estimated SHORT contracts held by sub-threshold traders across all months. "
     "Derived as: Open Interest minus Total Reportable Shorts."),
    ("nonrept_positions_long_old",   "int64", "Non-Reportable Positions", "Small Trader Long — Front Month",
     "Non-reportable LONG positions in the front-month contract only."),
    ("nonrept_positions_short_old",  "int64", "Non-Reportable Positions", "Small Trader Short — Front Month",
     "Non-reportable SHORT positions in the front-month contract only."),

    # Pct of OI
    ("pct_of_oi_noncomm_long_all",    "float64", "Percentage of Open Interest", "% OI — Spec Long (All)",
     "Non-commercial long positions as a % of total open interest (all months)."),
    ("pct_of_oi_noncomm_short_all",   "float64", "Percentage of Open Interest", "% OI — Spec Short (All)",
     "Non-commercial short positions as a % of total open interest (all months)."),
    ("pct_of_oi_noncomm_spread",      "float64", "Percentage of Open Interest", "% OI — Spec Spread (All)",
     "Non-commercial spread positions as a % of total open interest (all months)."),
    ("pct_of_oi_comm_long_all",       "float64", "Percentage of Open Interest", "% OI — Hedger Long (All)",
     "Commercial long positions as a % of total open interest (all months)."),
    ("pct_of_oi_comm_short_all",      "float64", "Percentage of Open Interest", "% OI — Hedger Short (All)",
     "Commercial short positions as a % of total open interest (all months)."),
    ("pct_of_oi_tot_rept_long_all",   "float64", "Percentage of Open Interest", "% OI — Total Rept Long (All)",
     "Total reportable long positions as a % of total open interest (all months)."),
    ("pct_of_oi_tot_rept_short",      "float64", "Percentage of Open Interest", "% OI — Total Rept Short (All)",
     "Total reportable short positions as a % of total open interest (all months)."),
    ("pct_of_oi_nonrept_long_all",    "float64", "Percentage of Open Interest", "% OI — Small Long (All)",
     "Non-reportable long positions as a % of total open interest (all months)."),
    ("pct_of_oi_nonrept_short_all",   "float64", "Percentage of Open Interest", "% OI — Small Short (All)",
     "Non-reportable short positions as a % of total open interest (all months)."),
    ("pct_of_oi_noncomm_long_old",    "float64", "Percentage of Open Interest", "% OI — Spec Long (Front)",
     "Non-commercial long positions as a % of front-month open interest only."),
    ("pct_of_oi_noncomm_short_old",   "float64", "Percentage of Open Interest", "% OI — Spec Short (Front)",
     "Non-commercial short positions as a % of front-month open interest only."),
    ("pct_of_oi_noncomm_spread_1",    "float64", "Percentage of Open Interest", "% OI — Spec Spread (Front)",
     "Non-commercial spread positions as a % of front-month open interest only."),
    ("pct_of_oi_comm_long_old",       "float64", "Percentage of Open Interest", "% OI — Hedger Long (Front)",
     "Commercial long positions as a % of front-month open interest only."),
    ("pct_of_oi_comm_short_old",      "float64", "Percentage of Open Interest", "% OI — Hedger Short (Front)",
     "Commercial short positions as a % of front-month open interest only."),
    ("pct_of_oi_tot_rept_long_old",   "float64", "Percentage of Open Interest", "% OI — Total Rept Long (Front)",
     "Total reportable long positions as a % of front-month open interest only."),
    ("pct_of_oi_tot_rept_short_1",    "float64", "Percentage of Open Interest", "% OI — Total Rept Short (Front)",
     "Total reportable short positions as a % of front-month open interest only."),
    ("pct_of_oi_nonrept_long_old",    "float64", "Percentage of Open Interest", "% OI — Small Long (Front)",
     "Non-reportable long positions as a % of front-month open interest only."),
    ("pct_of_oi_nonrept_short_old",   "float64", "Percentage of Open Interest", "% OI — Small Short (Front)",
     "Non-reportable short positions as a % of front-month open interest only."),
    ("pct_of_open_interest_other",    "float64", "Percentage of Open Interest", "% OI — Deferred Months",
     "Percentage of total open interest that resides in deferred (non-front-month) contracts. "
     "Computed as: (Open Interest All − Open Interest Old) / Open Interest All × 100."),

    # Traders
    ("traders_tot_all",           "int64", "Number of Traders", "# Traders Total (All)",
     "Total count of distinct reportable trading entities across all delivery months."),
    ("traders_noncomm_long_all",  "int64", "Number of Traders", "# Spec Longs (All)",
     "Number of non-commercial traders holding a net long position, all months. "
     "Note: one trader can appear in multiple subcategories, so counts do not sum to total."),
    ("traders_noncomm_short_all", "int64", "Number of Traders", "# Spec Shorts (All)",
     "Number of non-commercial traders holding a net short position, all months."),
    ("traders_noncomm_spread_all","int64", "Number of Traders", "# Spec Spreads (All)",
     "Number of non-commercial traders with active spread positions, all months."),
    ("traders_comm_long_all",     "int64", "Number of Traders", "# Hedger Longs (All)",
     "Number of commercial traders net long, all months."),
    ("traders_comm_short_all",    "int64", "Number of Traders", "# Hedger Shorts (All)",
     "Number of commercial traders net short, all months."),
    ("traders_tot_rept_long_all", "int64", "Number of Traders", "# Total Rept Long (All)",
     "Number of all reportable traders net long, all months."),
    ("traders_tot_rept_short_all","int64", "Number of Traders", "# Total Rept Short (All)",
     "Number of all reportable traders net short, all months."),
    ("traders_tot_old",           "int64", "Number of Traders", "# Traders Total (Front)",
     "Total count of distinct reportable trading entities in front-month contracts only."),
    ("traders_noncomm_long_old",  "int64", "Number of Traders", "# Spec Longs (Front)",
     "Number of non-commercial traders net long in the front-month contract."),
    ("traders_noncomm_short_old", "int64", "Number of Traders", "# Spec Shorts (Front)",
     "Number of non-commercial traders net short in the front-month contract."),
    ("traders_noncomm_spead_old", "int64", "Number of Traders", "# Spec Spreads (Front)",
     "Number of non-commercial spread traders in front-month. "
     "(Note: 'spead' is a typo in the original CFTC source data — should be 'spread'.)"),
    ("traders_comm_long_old",     "int64", "Number of Traders", "# Hedger Longs (Front)",
     "Number of commercial traders net long in front-month."),
    ("traders_comm_short_old",    "int64", "Number of Traders", "# Hedger Shorts (Front)",
     "Number of commercial traders net short in front-month."),
    ("traders_tot_rept_long_old", "int64", "Number of Traders", "# Total Rept Long (Front)",
     "Number of all reportable traders net long in front-month."),
    ("traders_tot_rept_short_old","int64", "Number of Traders", "# Total Rept Short (Front)",
     "Number of all reportable traders net short in front-month."),

    # Concentration
    ("conc_gross_le_4_tdr_long",      "float64", "Concentration Ratios", "Conc. Gross Long — Top 4 (All)",
     "% of total open interest held LONG (gross) by the 4 largest reportable traders, all months. "
     "'Gross' means long and short are counted independently before aggregating."),
    ("conc_gross_le_4_tdr_short",     "float64", "Concentration Ratios", "Conc. Gross Short — Top 4 (All)",
     "% of total open interest held SHORT (gross) by the 4 largest reportable traders, all months."),
    ("conc_gross_le_8_tdr_long",      "float64", "Concentration Ratios", "Conc. Gross Long — Top 8 (All)",
     "% of total open interest held LONG (gross) by the 8 largest reportable traders, all months."),
    ("conc_gross_le_8_tdr_short",     "float64", "Concentration Ratios", "Conc. Gross Short — Top 8 (All)",
     "% of total open interest held SHORT (gross) by the 8 largest reportable traders, all months."),
    ("conc_net_le_4_tdr_long_all",    "float64", "Concentration Ratios", "Conc. Net Long — Top 4 (All)",
     "% of total open interest held NET LONG by the 4 largest traders, all months. "
     "'Net' = each trader's long minus short before the group is aggregated."),
    ("conc_net_le_4_tdr_short_all",   "float64", "Concentration Ratios", "Conc. Net Short — Top 4 (All)",
     "% of total open interest held NET SHORT by the 4 largest traders, all months."),
    ("conc_net_le_8_tdr_long_all",    "float64", "Concentration Ratios", "Conc. Net Long — Top 8 (All)",
     "% of total open interest held NET LONG by the 8 largest traders, all months."),
    ("conc_net_le_8_tdr_short_all",   "float64", "Concentration Ratios", "Conc. Net Short — Top 8 (All)",
     "% of total open interest held NET SHORT by the 8 largest traders, all months."),
    ("conc_gross_le_4_tdr_long_1",    "float64", "Concentration Ratios", "Conc. Gross Long — Top 4 (Front)",
     "Gross long concentration for top 4 traders in the front-month contract only."),
    ("conc_gross_le_4_tdr_short_1",   "float64", "Concentration Ratios", "Conc. Gross Short — Top 4 (Front)",
     "Gross short concentration for top 4 traders in the front-month contract only."),
    ("conc_gross_le_8_tdr_long_1",    "float64", "Concentration Ratios", "Conc. Gross Long — Top 8 (Front)",
     "Gross long concentration for top 8 traders in the front-month contract only."),
    ("conc_gross_le_8_tdr_short_1",   "float64", "Concentration Ratios", "Conc. Gross Short — Top 8 (Front)",
     "Gross short concentration for top 8 traders in the front-month contract only."),
    ("conc_net_le_4_tdr_long_old",    "float64", "Concentration Ratios", "Conc. Net Long — Top 4 (Front)",
     "Net long concentration for top 4 traders in front-month only."),
    ("conc_net_le_4_tdr_short_old",   "float64", "Concentration Ratios", "Conc. Net Short — Top 4 (Front)",
     "Net short concentration for top 4 traders in front-month only."),
    ("conc_net_le_8_tdr_long_old",    "float64", "Concentration Ratios", "Conc. Net Long — Top 8 (Front)",
     "Net long concentration for top 8 traders in front-month only."),
    ("conc_net_le_8_tdr_short_old",   "float64", "Concentration Ratios", "Conc. Net Short — Top 8 (Front)",
     "Net short concentration for top 8 traders in front-month only."),

    # Changes
    ("change_in_open_interest_all",   "float64", "Week-over-Week Changes", "Δ Open Interest",
     "Change in total open interest (all months) vs. the prior Tuesday's report."),
    ("change_in_noncomm_long_all",    "float64", "Week-over-Week Changes", "Δ Spec Long",
     "Change in non-commercial long positions vs. prior week. "
     "Positive = speculators added longs; negative = reduced or flipped."),
    ("change_in_noncomm_short_all",   "float64", "Week-over-Week Changes", "Δ Spec Short",
     "Change in non-commercial short positions vs. prior week."),
    ("change_in_noncomm_spead_all",   "float64", "Week-over-Week Changes", "Δ Spec Spread",
     "Change in non-commercial spread positions vs. prior week. "
     "(Note: 'spead' is a typo in the original CFTC source data.)"),
    ("change_in_comm_long_all",       "float64", "Week-over-Week Changes", "Δ Hedger Long",
     "Change in commercial long positions vs. prior week."),
    ("change_in_comm_short_all",      "float64", "Week-over-Week Changes", "Δ Hedger Short",
     "Change in commercial short positions vs. prior week."),
    ("change_in_tot_rept_long_all",   "float64", "Week-over-Week Changes", "Δ Total Rept Long",
     "Change in total reportable long positions vs. prior week."),
    ("change_in_tot_rept_short",      "float64", "Week-over-Week Changes", "Δ Total Rept Short",
     "Change in total reportable short positions vs. prior week."),
    ("change_in_nonrept_long_all",    "float64", "Week-over-Week Changes", "Δ Small Long",
     "Change in non-reportable (small trader) long positions vs. prior week."),
    ("change_in_nonrept_short_all",   "float64", "Week-over-Week Changes", "Δ Small Short",
     "Change in non-reportable short positions vs. prior week."),
]

# ── Write data rows ──────────────────────────────────────────────────────────
current_section = None
data_start_row = 4

for row_idx, (col_name, dtype, section, short_label, description) in enumerate(rows, start=data_start_row):
    sec_header_color, sec_fill_color = SECTION_COLOURS.get(section, (GREY_HEADER, LIGHT_GREY))

    # Insert a section-header row when the section changes
    if section != current_section:
        sec_row = row_idx + (row_idx - data_start_row)  # shift for inserted headers
        # recalculate actual excel row accounting for already-inserted headers

    # Use alternating row shading within each section
    bg = WHITE if row_idx % 2 == 0 else LIGHT_GREY

    values = [col_name, dtype, section, short_label, description]
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = cell_font(bold=(col_idx == 1), size=10,
                              color="1F3864" if col_idx == 1 else "000000")
        cell.fill = fill(bg)
        cell.border = thin_border
        cell.alignment = wrap_align("left")

ws.row_dimensions[row_idx].height = 15

# Re-write with proper section grouping ──────────────────────────────────────
# Clear and redo from row 4 with section banners
for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
    for cell in row:
        cell.value = None

excel_row = 4
current_section = None

for (col_name, dtype, section, short_label, description) in rows:
    sec_header_color, sec_fill_color = SECTION_COLOURS.get(section, (GREY_HEADER, LIGHT_GREY))

    # Section banner
    if section != current_section:
        ws.merge_cells(f"A{excel_row}:E{excel_row}")
        banner = ws.cell(row=excel_row, column=1, value=f"  {section}")
        banner.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
        banner.fill = fill(sec_header_color)
        banner.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        banner.border = Border(
            left=Side(style="medium", color=sec_header_color),
            right=Side(style="medium", color=sec_header_color),
            top=Side(style="medium", color=sec_header_color),
            bottom=Side(style="medium", color=sec_header_color),
        )
        ws.row_dimensions[excel_row].height = 22
        excel_row += 1
        current_section = section

    # Alternating row shading
    row_in_section = excel_row % 2
    bg = WHITE if row_in_section == 0 else sec_fill_color

    values = [col_name, dtype, section, short_label, description]
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=excel_row, column=col_idx, value=val)
        cell.fill = fill(bg)
        cell.border = thin_border
        cell.alignment = wrap_align("left")
        if col_idx == 1:
            cell.font = Font(name="Courier New", bold=True, size=9, color=sec_header_color)
        elif col_idx == 2:
            cell.font = Font(name="Courier New", size=9, color="505050", italic=True)
        elif col_idx == 4:
            cell.font = Font(name="Calibri", bold=True, size=10, color="000000")
        else:
            cell.font = cell_font(size=10)

    ws.row_dimensions[excel_row].height = 42
    excel_row += 1

# ── Column widths ────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 38   # Column Name
ws.column_dimensions["B"].width = 12   # Data Type
ws.column_dimensions["C"].width = 28   # Section
ws.column_dimensions["D"].width = 32   # Short Label
ws.column_dimensions["E"].width = 72   # Full Description

# ── Freeze panes ─────────────────────────────────────────────────────────────
ws.freeze_panes = "A4"

# ── Auto-filter ──────────────────────────────────────────────────────────────
ws.auto_filter.ref = f"A3:E3"

# ── Add a Glossary sheet ─────────────────────────────────────────────────────
gs = wb.create_sheet("Glossary")
gs.sheet_view.showGridLines = False

glossary_title = gs.merge_cells("A1:B1")
t = gs["A1"]
t.value = "Key Concepts Glossary"
t.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
t.fill = fill(DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
gs.row_dimensions[1].height = 26

for col, hdr in enumerate(["Term", "Definition"], start=1):
    c = gs.cell(row=2, column=col, value=hdr)
    c.font = header_font(size=10)
    c.fill = fill(GREY_HEADER)
    c.alignment = wrap_align("center")
    c.border = thin_border
gs.row_dimensions[2].height = 18

glossary = [
    ("Open Interest",
     "The total number of outstanding futures contracts that have not been settled. "
     "Each contract has a buyer and a seller; open interest counts one side. "
     "Rising OI alongside rising price typically confirms a trend; falling OI suggests unwinding."),
    ("Commercial Trader (Hedger)",
     "An entity that uses futures to hedge a real-world business exposure (e.g. a pension fund "
     "hedging equity risk). They register a hedge exemption with the CFTC. Often called 'smart money' "
     "because their trading is driven by fundamental risk management rather than speculation."),
    ("Non-Commercial Trader (Speculator)",
     "An entity trading futures for profit rather than to hedge an underlying exposure. "
     "Includes hedge funds, CTAs, proprietary trading desks, and large retail participants. "
     "Their collective net position is widely watched as a sentiment/positioning indicator."),
    ("Non-Reportable Trader (Small Trader)",
     "Any trader whose position is below the CFTC's minimum reporting threshold. "
     "Their aggregate position is implied (Open Interest minus all reportable positions). "
     "Typically retail or small institutional participants."),
    ("Reportable Threshold",
     "The minimum position size (set per contract by the CFTC) above which a trader must file "
     "daily position reports. For S&P 500 futures this is typically 1,000 contracts net on either side."),
    ("Spreading",
     "A strategy where a trader simultaneously holds a LONG in one delivery month and a SHORT in "
     "a different delivery month of the same (or related) contract. The CFTC separates these "
     "from outright long/short counts because they are not directional bets on market level."),
    ("Old Futures (Front Month)",
     "Positions in the nearest-expiry (nearby) delivery month. The 'old' designation refers to "
     "the most active contract that is closest to expiration. Useful for studying roll dynamics."),
    ("Other Futures (Deferred Months)",
     "Positions in all delivery months BEYOND the front month. Not directly in the column names "
     "but implied as (All − Old). Large deferred OI can indicate hedging demand further out the curve."),
    ("Gross Position",
     "Longs and shorts counted independently before netting. Used in the concentration ratio "
     "calculations. Gross long + Gross short > Net position in absolute terms."),
    ("Net Position",
     "A trader's long contracts minus their short contracts. Net long = more longs than shorts; "
     "net short = more shorts than longs. Used in concentration ratios and sentiment indicators."),
    ("Concentration Ratio (Top 4 / Top 8)",
     "The percentage of total open interest controlled by the 4 or 8 largest reportable traders. "
     "A high concentration ratio signals a few dominant players — useful for detecting crowding "
     "or potential for large forced unwinds."),
    ("Net Speculator Position (Derived)",
     "noncomm_long_all − noncomm_short_all. A widely used sentiment gauge. Extreme net longs "
     "historically precede corrections; extreme net shorts can precede short squeezes."),
    ("Commercial Net (Derived)",
     "comm_long_all − comm_short_all. Hedgers are often treated as contra-indicators: "
     "extreme net short by commercials may mean they are heavily hedged (bearish on owned asset); "
     "extreme net long may signal bullish underlying demand."),
]

for r_idx, (term, defn) in enumerate(glossary, start=3):
    bg = WHITE if r_idx % 2 == 0 else LIGHT_BLUE
    tc = gs.cell(row=r_idx, column=1, value=term)
    tc.font = Font(name="Calibri", bold=True, size=10, color=DARK_BLUE)
    tc.fill = fill(bg)
    tc.border = thin_border
    tc.alignment = wrap_align()

    dc = gs.cell(row=r_idx, column=2, value=defn)
    dc.font = cell_font(size=10)
    dc.fill = fill(bg)
    dc.border = thin_border
    dc.alignment = wrap_align()
    gs.row_dimensions[r_idx].height = 52

gs.column_dimensions["A"].width = 36
gs.column_dimensions["B"].width = 90
gs.freeze_panes = "A3"

# ── Save ──────────────────────────────────────────────────────────────────────
output_path = r"c:\ProjectsDannyDavis\VolatilityPredictionProject\DataSourceExploration\COT_Column_Reference.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
